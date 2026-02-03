from telethon import TelegramClient
import asyncio
import os
import sqlite3
from dotenv import load_dotenv
from datetime import timezone
from openpyxl import Workbook, load_workbook

# ---------- CONFIG ----------
DB_FILE = "cumta_alerts.db"
CHANNEL = "@CumtaAlertsEnglishChannel"
OUTPUT_XLSX = "alerts_missing_alert_type_with_previous.xlsx"
UPDATE_XLSX = "update_type.xlsx"

load_dotenv()
API_ID = os.getenv("API_ID")
API_HASH = os.getenv("API_HASH")


# ---------- DB ----------
def get_missing_alert_ids():
    """Get IDs of messages where alert_type is missing"""
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute(
        """
        SELECT DISTINCT raw_message_id
        FROM alerts
        WHERE alert_type IS NULL
           OR TRIM(alert_type) = ''
        """
    )
    ids = {row[0] for row in cur.fetchall()}
    conn.close()
    return ids


def update_alert_type_from_excel(xlsx_file):
    """Update alert_type column in DB from Excel"""
    wb = load_workbook(xlsx_file)
    ws = wb.active
    updated, skipped, unmatched = 0, 0, 0

    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()

    # Assuming first row is headers
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        msg_id, alert_type = row
        if not alert_type or not msg_id:
            skipped += 1
            continue
        cur.execute("SELECT COUNT(*) FROM alerts WHERE raw_message_id = ?", (msg_id,))
        count = cur.fetchone()[0]
        if count == 0:
            unmatched += 1
            continue
        cur.execute(
            "UPDATE alerts SET alert_type = ? WHERE raw_message_id = ?",
            (alert_type.strip(), msg_id),
        )
        updated += 1

    conn.commit()
    conn.close()

    print("----- RESULT -----")
    print(f"Updated msgs   : {updated}")
    print(f"Skipped msgs   : {skipped}")
    print(f"Unmatched msgs : {unmatched}")


# ---------- TELEGRAM EXPORT ----------
async def export_missing_alerts():
    """Export messages with missing alert_type plus previous and pre-previous"""
    missing_ids = get_missing_alert_ids()
    if not missing_ids:
        print("No missing alert_type rows found.")
        return

    expanded_ids = set(missing_ids)
    for mid in missing_ids:
        if mid and mid > 1:
            expanded_ids.add(mid - 1)  # previous
        if mid and mid > 2:
            expanded_ids.add(mid - 2)  # pre-previous

    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Telegram Messages"
    ws.append(["telegram_message_id", "relation", "telegram_datetime_utc", "raw_text"])

    async with TelegramClient("cumta_alerttype_inspect", API_ID, API_HASH) as client:
        print(f"Connected. Fetching {len(expanded_ids)} messages from Telegram...\n")
        for msg_id in sorted(expanded_ids):
            try:
                msg = await client.get_messages(CHANNEL, ids=msg_id)
                if not msg:
                    continue
                if msg_id in missing_ids:
                    relation = "TARGET"
                elif msg_id == min(expanded_ids & {msg_id + 1, msg_id + 2}):
                    relation = "PREVIOUS"
                else:
                    relation = "PRE-PREVIOUS"

                ws.append(
                    [
                        msg.id,
                        relation,
                        (
                            msg.date.astimezone(timezone.utc).isoformat()
                            if msg.date
                            else None
                        ),
                        msg.text,
                    ]
                )
            except Exception as e:
                ws.append([msg_id, "ERROR", None, str(e)])

    wb.save(OUTPUT_XLSX)
    print(f"Saved inspection file to {OUTPUT_XLSX}")


# ---------- MAIN ----------
if __name__ == "__main__":
    if os.path.exists(UPDATE_XLSX):
        print(f"Found {UPDATE_XLSX}, updating DB...")
        update_alert_type_from_excel(UPDATE_XLSX)
    else:
        print(f"{UPDATE_XLSX} not found, exporting messages with missing alert_type...")
        asyncio.run(export_missing_alerts())
